from openpyxl import Workbook, load_workbook

wb = load_workbook('python learn-1\openpyxl learn\excel.xlsx')
ws = wb.active
print(ws['B3'].value)

ws['B3'].value = '98'

wb.save('python learn-1\openpyxl learn\excel.xlsx')

print(ws['B3'].value)